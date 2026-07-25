"""
Hotel room-block audit list generator.

CCA guarantees hotels a room block and pays for shortfall below the
contracted fill rate. Players book through third-party sites (Priceline,
Expedia) instead of the block link, so the hotel undercounts CCA-attributable
rooms. The organizer may demand an audit: he supplies a name list, the hotel
matches it against reservation folios, and every matched room-night comes off
the shortfall bill.

This tool builds that name list. Sources, best combined:

  - The public chessaction.com entry list for the event (scraped): every
    active + withdrawn player with full name, USCF ID, state, section.
    Available any time, including before entries close.
  - The organizer's registration-admin CSV export (optional --payers):
    adds home city/zip, side-event players not on the main entry-list
    page, and — critically — who PAID each entry. A parent who paid the
    entry fee likely booked the room under their own name, so external
    payers get their own row on the list.

Output is a two-sheet .xlsx:
  - "Audit List": clean LastName/FirstName/City/State/Zip — hand to the hotel.
  - "Reference":  per-person type (Player/Payer), data source, section,
    paid-for cross-reference, and data-quality flags — for the organizer.

Usage:
    python3 hotel_audit.py "southern open"                    # scrape only
    python3 hotel_audit.py "world open" --payers export.csv   # scrape + export
    python3 hotel_audit.py --payers export.csv                # export only
    python3 hotel_audit.py "southern open" -o audit.xlsx --year 2026

The admin export must have columns:
    LastName, FirstName, City, State, ZipCode, PayerName
with PayerName formatted "Last, First".
"""

import argparse
import csv
import logging
import re
import sys
from collections import Counter
from datetime import date
from pathlib import Path

from bs4 import BeautifulSoup

from scrape_entries import ENTRY_LIST_URL_TEMPLATE, _derive_entry_list_code
from scraper_utils import polite_session, respectful_get

logger = logging.getLogger(__name__)

# A payer covering more than this many players is registration staff
# processing mail/phone entries (observed: CCA staff paying 27-40 entries),
# not a parent booking a room. Kept on the list once, flagged, and their
# per-player cross-reference is suppressed as meaningless.
BULK_PAYER_THRESHOLD = 3

EXPORT_REQUIRED_COLUMNS = ("LastName", "FirstName", "PayerName")

AUDIT_SHEET_COLUMNS = ["LastName", "FirstName", "City", "State", "Zip"]
REFERENCE_SHEET_COLUMNS = ["LastName", "FirstName", "Type", "Source",
                           "Section", "Notes", "Flags"]

# Strips "(GM)", "(Withdrawn)" etc. Names on CCA lists never contain parens.
_PAREN_RE = re.compile(r"\([^)]*\)")


# ---------------------------------------------------------------------------
# Name normalization
# ---------------------------------------------------------------------------

def clean_name(s):
    """Strip parenthesized annotations ((GM), (Withdrawn)) and tidy spaces."""
    return " ".join(_PAREN_RE.sub("", s or "").split())


def smart_case(s):
    """ALL CAPS or all-lower -> Title Case; mixed case (McDonald) untouched."""
    s = (s or "").strip()
    if s.isupper() or s.islower():
        return s.title()
    return s


def name_key(last, first):
    """Identity key: (last name, FIRST TOKEN of first name), case-folded.

    First-token matching absorbs middle names: the entry list says
    "Brightwater, Ana Marie" while the admin export says "Brightwater, Ana" —
    same person. Validated against the 2026 World Open: all 1,108
    entry-list players matched their export row on this key.
    """
    last = " ".join(clean_name(last).lower().split())
    first_tokens = clean_name(first).lower().split()
    return (last, first_tokens[0] if first_tokens else "")


def split_last_first(name):
    """Split a "Last, First M" string; falls back for comma-less input."""
    name = clean_name(name)
    if "," in name:
        last, first = name.split(",", 1)
        return last.strip(), first.strip()
    parts = name.split()
    if len(parts) > 1:
        return parts[-1], " ".join(parts[:-1])
    return name, ""


# ---------------------------------------------------------------------------
# Entry-list scraping (public chessaction.com pages)
# ---------------------------------------------------------------------------

def resolve_event(event_name, year):
    """Resolve a free-text event name to its CCA entry-list URL."""
    code = _derive_entry_list_code(event_name)
    if not code:
        raise ValueError(f"cannot derive an entry-list code from {event_name!r}")
    yy = str(year)[-2:]
    return code, ENTRY_LIST_URL_TEMPLATE.format(code=code, yy=yy)


def fetch_entry_list(url):
    """GET the entry-list page politely; raise on non-200."""
    session = polite_session()
    resp = respectful_get(session, url)
    if resp.status_code != 200:
        raise RuntimeError(
            f"entry list fetch failed: HTTP {resp.status_code} for {url}\n"
            "Check the event name/year — the page may not be published yet."
        )
    return resp.text


def parse_entry_list(html):
    """Extract player rows from a CCA entry-list page.

    Page structure (see prediction/backend/scraper/cca.py blueprint):
    the player table has a header row with class="success" and player rows
    of >=8 <td> cells: Name | Name (alt) | FIDE ID | FIDE Rating | USCF ID |
    USCF Rating | State | Section | Schedule | Bye(s).

    The first cell carries the clean "Last, First M" in its data-order
    attribute; visible text may append "(GM)" titles and "(Withdrawn)".

    Returns a list of dicts:
        {last, first, uscf_id, state, section, withdrawn}
    """
    soup = BeautifulSoup(html, "html.parser")
    tables = soup.find_all("table")
    if not tables:
        return []

    player_table = None
    for t in tables:
        if t.find("tr", class_="success"):
            player_table = t
            break
    if player_table is None:
        player_table = max(tables, key=lambda t: len(t.find_all("tr")))

    players = []
    for row in player_table.find_all("tr"):
        if "success" in " ".join(row.get("class", [])):
            continue  # header row
        cols = row.find_all("td")
        if len(cols) < 8:
            continue
        cell_text = cols[0].get_text(strip=True)
        raw_name = cols[0].get("data-order") or cell_text
        name = clean_name(raw_name)
        if not name or name.lower().startswith(("player", "name", "filter")):
            continue
        last, first = split_last_first(name)
        players.append({
            "last": last,
            "first": first,
            "uscf_id": cols[4].get_text(strip=True) or None,
            "state": cols[6].get_text(strip=True) or "",
            "section": cols[7].get_text(strip=True) or "",
            "withdrawn": "(withdrawn)" in cell_text.lower(),
        })
    return players


def dedup_scraped(players):
    """Collapse duplicate rows within one scrape.

    USCF ID is the reliable identity where present (distinguishes two
    players who share a name); rows without an ID fall back to the name
    key. First sighting wins; a withdrawn duplicate never un-withdraws
    an active row.
    """
    seen = {}
    for p in players:
        key = ("id", p["uscf_id"]) if p["uscf_id"] else ("name", name_key(p["last"], p["first"]))
        if key not in seen:
            seen[key] = p
        elif seen[key]["withdrawn"] and not p["withdrawn"]:
            seen[key] = p
    return list(seen.values())


# ---------------------------------------------------------------------------
# Admin export (organizer's registration system pull)
# ---------------------------------------------------------------------------

def load_export(path):
    """Read the admin CSV export; validate required columns."""
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        missing = [c for c in EXPORT_REQUIRED_COLUMNS
                   if c not in (reader.fieldnames or [])]
        if missing:
            raise ValueError(
                f"export is missing required column(s) {missing}; "
                f"found: {reader.fieldnames}"
            )
        return [dict(r) for r in reader]


def pad_zip(z):
    """Restore leading zeros Excel strips (2155 -> 02155). ZIP+4 untouched."""
    z = (z or "").strip()
    if z.isdigit() and 0 < len(z) < 5:
        return z.zfill(5)
    return z


# ---------------------------------------------------------------------------
# Merge: scrape + export -> audit people
# ---------------------------------------------------------------------------

def _new_person(last, first, *, type_, source):
    return {
        "last": smart_case(last), "first": smart_case(first),
        "city": "", "state": "", "zip": "",
        "type": type_, "source": source, "section": "",
        "withdrawn": False, "paid_for": [], "flags": [],
    }


def _flag(person, msg):
    if msg not in person["flags"]:
        person["flags"].append(msg)


def _apply_address(person, row):
    """Copy the export row's address onto a person. The export is the
    address authority: the folio shows the guest's HOME address, and the
    entry list's State column is the USCF federation state, which can
    legitimately differ (live in NJ, play under NY)."""
    city = (row.get("City") or "").strip()
    state = (row.get("State") or "").strip()
    zip_ = (row.get("ZipCode") or "").strip()
    if city:
        person["city"] = smart_case(city)
    if state:
        person["state"] = state.upper()
    if zip_:
        person["zip"] = pad_zip(zip_)
    if "," in city:
        _flag(person, "check city field (looks like a full address)")


def build_people(scraped, export_rows):
    """Merge deduped entry-list players with the admin export.

    Returns (people, stats). Each person dict:
        {last, first, city, state, zip, type, source, section,
         withdrawn, paid_for, flags}

    Rules:
      - Cross-source identity is name_key() alone (validated: all 1,108
        WO26 entry-list players matched their export row on it). The
        entry list is the name (fuller, middle names) + section +
        withdrawn authority; the export is the address + payer authority.
      - Within the export, the same name in multiple states stays split:
        possibly two different people, and an extra row costs the audit
        nothing.
      - Every external payer (paid an entry, not a player themselves) is
        added once. Bulk payers (> BULK_PAYER_THRESHOLD players) are
        staff processing entries: flagged, cross-reference suppressed.
      - Nobody is silently dropped; data problems become flags.
    """
    stats = Counter()
    people = []

    # -- A. entry-list players: group by name key; distinct USCF IDs under
    #       one name are genuinely different people and keep separate rows.
    scrape_groups = {}   # name_key -> [person, ...]
    for p in scraped:
        stats["scraped_players"] += 1
        stats["scraped_withdrawn"] += p["withdrawn"]
        person = _new_person(p["last"], p["first"],
                             type_="Player", source="entry list")
        person["state"] = (p["state"] or "").upper()
        person["section"] = p["section"]
        person["withdrawn"] = p["withdrawn"]
        group = scrape_groups.setdefault(name_key(p["last"], p["first"]), [])
        group.append(person)
        people.append(person)
    for group in scrape_groups.values():
        if len(group) > 1:
            for person in group:
                _flag(person, "shares a name with another entry "
                              "(distinct USCF IDs); verify identity")

    # -- B. export rows: dedup players by name key, splitting on
    #       conflicting in-export states.
    export_groups = {}   # name_key -> {state_lower: (person-dict, row)}
    payer_of = {}        # payer name_key -> {"last","first","players","row"}
    for r in export_rows:
        stats["export_rows"] += 1
        last = (r.get("LastName") or "").strip()
        first = (r.get("FirstName") or "").strip()
        nkey = name_key(last, first)
        variants = export_groups.setdefault(nkey, {})
        state_l = (r.get("State") or "").strip().lower()
        # a stateless row folds into any existing variant; a stated row
        # claims the stateless variant if that is all that exists
        if state_l in variants:
            pass
        elif not state_l and variants:
            state_l = next(iter(variants))
        elif state_l and list(variants) == [""]:
            variants[state_l] = variants.pop("")
        if state_l not in variants:
            variants[state_l] = ({"last": last, "first": first}, r)

        payer = (r.get("PayerName") or "").strip()
        if payer:
            p_last, p_first = split_last_first(payer)
            pkey = name_key(p_last, p_first)
            if pkey != nkey:
                entry = payer_of.setdefault(
                    pkey, {"last": p_last, "first": p_first,
                           "players": set(), "row": r})
                entry["players"].add(f"{first} {last}".strip())

    # -- C. join: export enriches matched scrape rows; unmatched export
    #       variants become their own people (side events etc.).
    for nkey, variants in export_groups.items():
        multi_state = len(variants) > 1
        matched = scrape_groups.get(nkey)
        for state_l, (names, row) in variants.items():
            target = None
            if matched:
                # prefer the scrape row whose USCF state matches; else first
                target = next((p for p in matched
                               if p["state"].lower() == state_l), None)
                if target is None and not multi_state:
                    target = matched[0]
            if target is not None:
                stats["matched_both"] += 1
                target["source"] = "entry list+export"
            else:
                target = _new_person(names["last"], names["first"],
                                     type_="Player", source="export")
                stats["export_only"] += 1
                people.append(target)
            _apply_address(target, row)
            if not names["last"] or not names["first"]:
                _flag(target, "blank name in source data")
            if multi_state:
                _flag(target, "same name in multiple states; "
                              "possibly distinct people")

    if export_rows:
        stats["scrape_only"] = sum(
            1 for nkey, group in scrape_groups.items()
            if nkey not in export_groups for _ in group)

    # -- D. external payers (likely booked the room under their own name) --
    known_player_keys = set(scrape_groups) | set(export_groups)
    for pkey, entry in payer_of.items():
        if pkey in known_player_keys:
            stats["payers_already_players"] += 1
            continue
        stats["payers_added"] += 1
        person = _new_person(entry["last"], entry["first"],
                             type_="Payer", source="export")
        # a parent shares the player's home address — the folio match hook
        _apply_address(person, entry["row"])
        n = len(entry["players"])
        if n > BULK_PAYER_THRESHOLD:
            stats["bulk_payers"] += 1
            _flag(person, f"bulk payer, paid {n} entries "
                          "(likely registration staff)")
        else:
            person["paid_for"] = sorted(entry["players"])
        people.append(person)

    stats["final_people"] = len(people)
    return people, stats


# ---------------------------------------------------------------------------
# Workbook output
# ---------------------------------------------------------------------------

def write_workbook(people, out_path):
    """Two sheets: clean 'Audit List' for the hotel, 'Reference' for the
    organizer. Both sorted by last, first."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ordered = sorted(people, key=lambda p: (p["last"].lower(),
                                            p["first"].lower()))
    wb = openpyxl.Workbook()

    audit = wb.active
    audit.title = "Audit List"
    audit.append(AUDIT_SHEET_COLUMNS)
    for p in ordered:
        audit.append([p["last"], p["first"], p["city"], p["state"], p["zip"]])

    ref = wb.create_sheet("Reference")
    ref.append(REFERENCE_SHEET_COLUMNS)
    for p in ordered:
        notes = "Paid for: " + ", ".join(p["paid_for"]) if p["paid_for"] else ""
        flags = list(p["flags"])
        if p["withdrawn"]:
            flags.insert(0, "withdrawn from event; may still have booked")
        ref.append([p["last"], p["first"], p["type"], p["source"],
                    p["section"], notes, "; ".join(flags)])

    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    for ws, widths in ((audit, [18, 18, 22, 7, 11]),
                       (ref, [18, 18, 8, 18, 14, 40, 44])):
        for c in range(1, len(widths) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(vertical="center")
            ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(widths))}{ws.max_row}")

    wb.save(out_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Build a hotel room-block audit list for a CCA event.")
    parser.add_argument("event", nargs="?",
                        help='event name to scrape, e.g. "southern open" '
                             "(omit to use only --payers)")
    parser.add_argument("--payers", metavar="CSV",
                        help="registration-admin export CSV "
                             "(LastName,FirstName,City,State,ZipCode,PayerName)")
    parser.add_argument("-o", "--output", metavar="XLSX",
                        help="output path (default: hotel_audit_<CODE><YY>.xlsx)")
    parser.add_argument("--year", type=int, default=date.today().year,
                        help="event year (default: current year)")
    args = parser.parse_args(argv)

    if not args.event and not args.payers:
        parser.error("give an event name to scrape, a --payers export, or both")

    scraped, code = [], None
    if args.event:
        code, url = resolve_event(args.event, args.year)
        print(f"Fetching entry list: {url}")
        raw = parse_entry_list(fetch_entry_list(url))
        scraped = dedup_scraped(raw)
        if not scraped:
            raise RuntimeError(
                f"no players parsed from {url} — page layout may have "
                "changed (see structure_monitor.py) or the list is empty")
        print(f"  {len(scraped)} unique players "
              f"({sum(p['withdrawn'] for p in scraped)} withdrawn)")

    export_rows = []
    if args.payers:
        export_rows = load_export(args.payers)
        print(f"Export: {len(export_rows)} rows from {args.payers}")

    people, stats = build_people(scraped, export_rows)

    if args.output:
        out_path = Path(args.output)
    elif code:
        out_path = Path(f"hotel_audit_CCA_{code}{str(args.year)[-2:]}.xlsx")
    else:
        out_path = Path(args.payers).with_suffix("").with_name(
            Path(args.payers).stem + "_audit.xlsx")
    write_workbook(people, out_path)

    print()
    if export_rows and scraped:
        print(f"Matched in both sources : {stats['matched_both']}")
        print(f"Entry list only         : {stats['scrape_only']}")
        print(f"Export only (side events): {stats['export_only']}")
    if export_rows:
        print(f"External payers added   : {stats['payers_added']} "
              f"({stats['bulk_payers']} bulk/staff)")
        print(f"Payers already players  : {stats['payers_already_players']}")
    print(f"Final audit list        : {stats['final_people']} people")
    print(f"Written                 : {out_path}")
    return 0


if __name__ == "__main__":
    logging.basicConfig(level=logging.WARNING)
    sys.exit(main())
