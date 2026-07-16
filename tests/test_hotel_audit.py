"""
Tests for hotel_audit.py (hotel room-block audit list generator).

The HTML fixture is a trimmed verbatim capture of the real CCA_WO26 entry
list page (fetched 2026-07-15), so the parser is pinned to the live page
shape: section-summary table, class="success" header row, data-order name
attributes, "(GM)" titles, and "(IM)(Withdrawn)" annotations. A future
markup change fails here loudly instead of producing an empty audit list.

The admin-export fixture is synthetic (no real registrant data in the
repo): it mirrors the observed export shape, including a parent paying
for two siblings, ALL-CAPS payer spelling variants, a zip with a lost
leading zero, and a city field carrying a full address.
"""

import os
import sys

PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_DIR)

import pytest
from openpyxl import load_workbook

from hotel_audit import (
    build_people,
    dedup_scraped,
    load_export,
    name_key,
    pad_zip,
    parse_entry_list,
    resolve_event,
    smart_case,
    split_last_first,
    write_workbook,
)

FIXTURE_HTML = os.path.join(os.path.dirname(__file__), "fixtures",
                            "cca_entry_list_wo26.html")
FIXTURE_CSV = os.path.join(os.path.dirname(__file__), "fixtures",
                           "admin_export_sample.csv")


def _load_html():
    with open(FIXTURE_HTML, encoding="utf-8") as fh:
        return fh.read()


# ---------------------------------------------------------------------------
# Name helpers
# ---------------------------------------------------------------------------

def test_name_key_absorbs_middle_names_and_annotations():
    # entry list says "Brightwater, Ana Marie"; the export says "Brightwater, Ali"
    assert name_key("Brightwater", "Ana Marie") == name_key("BRIGHTWATER", "Ali")
    assert name_key("Adu (IM)(Withdrawn)", "Orlando") == ("adu", "orlando")
    assert name_key("", "") == ("", "")


def test_split_last_first():
    assert split_last_first("Zellwood, Jonquil") == ("Zellwood", "Jonquil")
    assert split_last_first("Sunderholm, Zephyr (GM)") == ("Sunderholm", "Zephyr")
    assert split_last_first("Merrick Jani") == ("Jani", "Merrick")
    assert split_last_first("Cher") == ("Cher", "")


def test_smart_case_preserves_mixed_case():
    assert smart_case("JANOWAY") == "Jani"
    assert smart_case("westmere") == "Westmere"
    assert smart_case("McDonald") == "McDonald"      # never mangled
    assert smart_case("Gianos-Steinberg") == "Gianos-Steinberg"


def test_pad_zip():
    assert pad_zip("2155") == "02155"    # Excel-stripped leading zero
    assert pad_zip("32608") == "32608"
    assert pad_zip("01748-3233") == "01748-3233"   # ZIP+4 untouched
    assert pad_zip("") == ""


# ---------------------------------------------------------------------------
# Entry-list parsing (real captured markup)
# ---------------------------------------------------------------------------

def test_parse_entry_list_real_markup():
    players = parse_entry_list(_load_html())
    assert len(players) == 5
    by_last = {p["last"]: p for p in players}

    # plain row: every field lands
    aaron = by_last["Aaron"]
    assert aaron["first"] == "Dennison"
    assert aaron["uscf_id"] == "12877422"
    assert aaron["state"] == "NY"
    assert aaron["section"] == "Open"
    assert aaron["withdrawn"] is False

    # middle name preserved (fuller than the admin export carries)
    assert by_last["Brightwater"]["first"] == "Ana Marie"

    # "(IM)(Withdrawn)" annotation: flagged withdrawn, name left clean
    adu = by_last["Adu"]
    assert adu["withdrawn"] is True
    assert "(" not in adu["first"] and "(" not in adu["last"]

    # "(GM)" title stripped
    assert by_last["Sunderholm"]["first"] == "Zephyr"
    assert by_last["Zellwood"]["first"] == "Jonquil"


def test_parse_entry_list_empty_page():
    assert parse_entry_list("<html><body>maintenance</body></html>") == []


def test_dedup_scraped_by_uscf_id():
    base = {"state": "NY", "section": "Open"}
    rows = [
        {"last": "Zellwood", "first": "Jonquil", "uscf_id": "15524414",
         "withdrawn": True, **base},
        {"last": "Zellwood", "first": "Jonquil", "uscf_id": "15524414",
         "withdrawn": False, **base},          # same person, active row wins
        {"last": "Zellwood", "first": "Jonquil", "uscf_id": "99999999",
         "withdrawn": False, **base},          # different person, same name
        {"last": "Lee", "first": "Justin", "uscf_id": None,
         "withdrawn": False, **base},
        {"last": "Lee", "first": "Justin M", "uscf_id": None,
         "withdrawn": False, **base},          # no ID: name-key fallback
    ]
    deduped = dedup_scraped(rows)
    assert len(deduped) == 3
    zellwoods = [p for p in deduped if p["last"] == "Zellwood"]
    assert len(zellwoods) == 2
    assert not any(p["withdrawn"] for p in zellwoods
                   if p["uscf_id"] == "15524414")


# ---------------------------------------------------------------------------
# Admin export loading
# ---------------------------------------------------------------------------

def test_load_export_fixture():
    rows = load_export(FIXTURE_CSV)
    assert len(rows) == 10
    assert rows[0]["LastName"] == "Aaron"


def test_load_export_missing_columns(tmp_path):
    bad = tmp_path / "bad.csv"
    bad.write_text("Name,City\nZellwood,Boston\n")
    with pytest.raises(ValueError, match="missing required column"):
        load_export(str(bad))


# ---------------------------------------------------------------------------
# Merge logic
# ---------------------------------------------------------------------------

def _scraped_sample():
    return dedup_scraped(parse_entry_list(_load_html()))


def test_build_people_scrape_only():
    people, stats = build_people(_scraped_sample(), [])
    assert stats["final_people"] == 5
    assert all(p["source"] == "entry list" for p in people)
    assert all(p["type"] == "Player" for p in people)
    adu = next(p for p in people if p["last"] == "Adu")
    assert adu["withdrawn"] is True


def test_build_people_merge_and_payers():
    export = load_export(FIXTURE_CSV)
    people, stats = build_people(_scraped_sample(), export)
    by_name = {(p["last"], p["first"]): p for p in people}

    # cross-source merge: fuller scrape name kept, export address attached,
    # even though the export state (home, FL) differs from the entry list's
    # USCF federation state (NJ)
    brightwater = by_name[("Brightwater", "Ana Marie")]
    assert brightwater["source"] == "entry list+export"
    assert brightwater["city"] == "Newark"
    assert brightwater["state"] == "FL"
    assert brightwater["zip"] == "07102"        # leading zero restored
    assert ("Brightwater", "Ali") not in by_name  # no duplicate from the export

    # siblings kept, parent payer added exactly once with cross-reference
    parent = by_name[("Quillon", "Rosalind")]
    assert parent["type"] == "Payer"
    assert parent["paid_for"] == ["Delphine Quillon", "Torrance Quillon"]
    assert parent["city"] == "Westmere"   # borrowed from the kids' rows

    # a payer who IS a player (Nina Lopez paid by player Zellwood) adds no row
    assert stats["payers_already_players"] == 1
    zellwood = by_name[("Zellwood", "Jonquil")]
    assert zellwood["type"] == "Player"

    # bulk payer (4 entries > threshold 3): flagged as staff,
    # cross-reference suppressed, address borrowed from first entry
    staff = by_name[("Carrier", "Robert")]
    assert staff["type"] == "Payer"
    assert staff["paid_for"] == []
    assert any("bulk payer" in f for f in staff["flags"])
    assert staff["city"] == "Reston"

    # export-only side-event player still on the list
    side = by_name[("Marbleton", "Marigold")]
    assert side["source"] == "export"

    # data-quality flags
    assert any("check city field" in f for f in side["flags"])

    # reconciliation: 5 scraped + 8 export-only players + 2 payers = 15
    assert stats["matched_both"] == 2
    assert stats["scrape_only"] == 3          # Adu, Sunderholm, Zellwood
    assert stats["export_only"] == 8
    assert stats["payers_added"] == 2
    assert stats["bulk_payers"] == 1
    assert stats["final_people"] == 15


def test_build_people_same_name_multiple_states_stays_split():
    export = [
        {"LastName": "Lee", "FirstName": "Justin", "City": "Chicago",
         "State": "IL", "ZipCode": "60601", "PayerName": "Lee, Justin"},
        {"LastName": "Lee", "FirstName": "Justin", "City": "Phoenix",
         "State": "AZ", "ZipCode": "85001", "PayerName": "Lee, Justin"},
    ]
    people, stats = build_people([], export)
    assert stats["final_people"] == 2
    assert all(any("multiple states" in f for f in p["flags"])
               for p in people)


def test_build_people_blank_name_flagged_not_dropped():
    export = [{"LastName": "Gianos-Steinberg", "FirstName": "",
               "City": "Cudahy", "State": "WI", "ZipCode": "53110",
               "PayerName": "Gianos-Steinberg, Alex"}]
    people, stats = build_people([], export)
    kid = next(p for p in people if p["type"] == "Player")
    assert any("blank name" in f for f in kid["flags"])
    # the named payer still gets a row: that is the folio match
    assert any(p["type"] == "Payer" and p["first"] == "Alex" for p in people)


def test_build_people_payer_spelling_variants_collapse():
    row = {"City": "Westmere", "State": "FL", "ZipCode": "32608"}
    export = [
        {"LastName": "Jani", "FirstName": "Aakash",
         "PayerName": "JANOWAY, MERRICK", **row},
        {"LastName": "Jani", "FirstName": "Suraj",
         "PayerName": "Jani, Merrick", **row},
    ]
    people, stats = build_people([], export)
    payers = [p for p in people if p["type"] == "Payer"]
    assert len(payers) == 1
    assert payers[0]["last"] == "Jani" and payers[0]["first"] == "Merrick"


# ---------------------------------------------------------------------------
# Workbook output
# ---------------------------------------------------------------------------

def test_write_workbook(tmp_path):
    export = load_export(FIXTURE_CSV)
    people, _ = build_people(_scraped_sample(), export)
    out = tmp_path / "audit.xlsx"
    write_workbook(people, out)

    wb = load_workbook(out)
    assert wb.sheetnames == ["Audit List", "Reference"]

    audit = wb["Audit List"]
    rows = list(audit.iter_rows(values_only=True))
    assert rows[0] == ("LastName", "FirstName", "City", "State", "Zip")
    assert len(rows) - 1 == len(people)
    # sorted by last name for folio matching
    lasts = [r[0].lower() for r in rows[1:]]
    assert lasts == sorted(lasts)

    ref = wb["Reference"]
    ref_rows = list(ref.iter_rows(values_only=True))
    assert ref_rows[0] == ("LastName", "FirstName", "Type", "Source",
                           "Section", "Notes", "Flags")
    notes = {(r[0], r[1]): (r[5] or "") for r in ref_rows[1:]}
    assert notes[("Quillon", "Rosalind")].startswith("Paid for: ")


# ---------------------------------------------------------------------------
# Event resolution
# ---------------------------------------------------------------------------

def test_resolve_event():
    code, url = resolve_event("southern open", 2026)
    assert code == "SO"
    assert url == ("https://www.chessaction.com/tournaments/advlists/CCA/"
                   "CCA_SO26/CCA_SO26_alp_n.html")
    code, url = resolve_event("2026 World Open", 2026)
    assert code == "WO"
    assert "CCA_WO26_alp_n.html" in url


def test_resolve_event_unknown_falls_back_to_initials():
    code, _ = resolve_event("Made Up Open", 2026)
    assert code == "MUO"
